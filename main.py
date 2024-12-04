import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

class ExcelProcessor:
    def __init__(self, input_file, sheet_name, municipio_nome):
        self.input_file = input_file
        self.sheet_name = sheet_name
        self.municipio_nome = municipio_nome
        self.df = None
        self.selected_columns = []

    def load_file(self):
        """Carregar o arquivo Excel com tratamento de erros."""
        try:
            # Especificando que os cabeçalhos estão na linha 2 (index 1)
            self.df = pd.read_excel(self.input_file, sheet_name=self.sheet_name, header=1)
            print("Arquivo carregado com sucesso.")
            print("Colunas encontradas:", self.df.columns)  # Imprime os nomes das colunas
        except FileNotFoundError:
            print("Arquivo não encontrado. Verifique o nome e o caminho.")
            exit()

    def clean_data(self):
        """Limpeza de dados."""
        if self.df is not None:
            # Normaliza os nomes das colunas para facilitar o acesso
            self.df.columns = [col.strip().upper() for col in self.df.columns]  # Remove espaços e converte para maiúsculas
            self.df.drop_duplicates(inplace=True)
            for col in self.df.columns:
                if pd.api.types.is_numeric_dtype(self.df[col]):
                    self.df[col] = self.df[col].fillna(0)
                elif pd.api.types.is_datetime64_any_dtype(self.df[col]):
                    self.df[col] = self.df[col].fillna(pd.NaT)
                else:
                    self.df[col] = self.df[col].fillna('')
            print("Dados limpos com sucesso.")
        else:
            print("DataFrame vazio. Carregue o arquivo primeiro.")
            exit()

    def select_columns(self):
        """Selecionar as colunas específicas para a exportação."""
        # Ajuste para corresponder aos nomes exatos das colunas carregadas
        self.selected_columns = ['NOME', 'CPF', 'LIQUIDO']
        print("Colunas selecionadas: NOME, CPF, LIQUIDO.")

    def export_to_excel(self, output_file):
        """Exportar os dados para um novo arquivo Excel com o formato solicitado."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório Líquido"

            # Título do Município (A1:D1 mesclado)
            ws.merge_cells('A1:D1')
            ws['A1'] = f"Relatório de {self.municipio_nome}"
            ws['A1'].font = Font(size=16, bold=True, color="000000")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A1'].fill = PatternFill(start_color="D9D9D6", end_color="D9D9D6", fill_type="solid")

            # Nome da Secretaria (A2)
            ws['A2'] = f"Secretaria: {self.sheet_name}"

            # Nome da Página (B2)
            ws['B2'] = f"Nome da Página: {self.sheet_name}"

            # Data e hora do relatório (C2)
            ws['C2'] = f"Data e Hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            # Cabeçalhos das colunas
            ws['A3'] = "Numeração"
            ws['B3'] = "Nome"
            ws['C3'] = "CPF"
            ws['D3'] = "Líquido"

            # Estilo dos cabeçalhos (letra preta e fundo cinza)
            bold_font = Font(bold=True, color="000000")  # Cor preta para o texto
            center_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
            title_fill = PatternFill(start_color="D9D9D6", end_color="D9D9D6", fill_type="solid")  # Fundo cinza

            # Aplicar estilos aos cabeçalhos
            for col_idx, col_name in enumerate(["Numeração", "Nome", "CPF", "Líquido"], start=1):
                cell = ws[f"{get_column_letter(col_idx)}3"]
                cell.value = col_name
                cell.font = bold_font
                cell.fill = title_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            # Escrever os dados nas células abaixo dos cabeçalhos, começando da linha 4
            for row_idx, row in enumerate(self.df[self.selected_columns].values, start=4):
                # Adicionar a numeração (1, 2, 3, ...)
                num_cell = ws.cell(row=row_idx, column=1, value=row_idx - 3)
                num_cell.alignment = center_alignment
                num_cell.border = thin_border  # Aplica borda à numeração
                for col_idx, value in enumerate(row, start=2):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = center_alignment
                    cell.border = thin_border

            # Inserir total no final
            total_row = len(self.df) + 4
            ws.cell(row=total_row, column=3, value="TOTAL")
            ws.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row-1})")
            ws[f"D{total_row}"].font = Font(bold=True)
            ws[f"D{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"D{total_row}"].border = thin_border
            ws[f"C{total_row}"].alignment = Alignment(horizontal="right", vertical="center")
            ws[f"C{total_row}"].border = thin_border

            # Ajustar a largura das colunas automaticamente
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_length = max(len(str(cell.value) if cell.value is not None else "") for cell in col_cells)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            # Salvar o arquivo com o nome do município e da secretaria
            output_file = f"relatorio_{self.sheet_name}.xlsx"
            wb.save(output_file)
            print(f"Relatório exportado para {output_file} com sucesso.")
        except Exception as e:
            print(f"Ocorreu um erro ao exportar para Excel: {e}")

# Uso da classe ExcelProcessor
def main():
    municipio_nome = 'NomeDoMunicipio'  # Exemplo de município
    processor = ExcelProcessor(input_file='FATURAMENTO - SAO MIGUEL DOS CAMPOS - 05 2024.xlsx', sheet_name='Geral', municipio_nome=municipio_nome)
    processor.load_file()
    processor.clean_data()
    processor.select_columns()
    processor.export_to_excel(output_file=f"relatorio_{municipio_nome}_caps.xlsx")

if __name__ == "__main__":
    main()
